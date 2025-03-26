import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Color, PatternFill
import sqlite3
from collections import defaultdict
from fractions import Fraction

align = Alignment(vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
fill = PatternFill(patternType='solid',
                   fill_type='solid',
                   fgColor=Color('FFFF00'))


instock_field_order = {'CATAGORY_NAME': 1, 'GOODS_NAME': 2, 'STOCK_DATE': 3, 'GOODS_NUM': 4,
                           'GOODS_UNIT': 5, 'GOODS_PRICE': 6, 'OP_AREA': 7, 'OP_PERSON': 8, 'TOTAL': 15}

outstock_field_order = {'CATAGORY_NAME': 1, 'GOODS_NAME': 2, 'STOCK_DATE': 9, 'GOODS_NUM': 10,
                           'GOODS_UNIT': 11, 'GOODS_PRICE': 12, 'OP_AREA': 13, 'OP_PERSON': 14, 'TOTAL': 16}

def get_total_in_out(goods_name):
    # 连接数据库
    con = sqlite3.connect('config/WarehouseManagement.db')
    cur = con.cursor()

    # 查询总入库数量
    sql_in = '''
    SELECT SUM(GOODS_NUM) FROM STOCK 
    JOIN GOODS ON STOCK.GOODS_ID = GOODS.ID 
    WHERE GOODS.GOODS_NAME =? AND STOCK.STOCK_TYPE = 1
    '''
    cur.execute(sql_in, (goods_name,))
    total_in = cur.fetchone()[0] or 0

    # 查询总出库数量
    sql_out = '''
    SELECT SUM(GOODS_NUM) FROM STOCK 
    JOIN GOODS ON STOCK.GOODS_ID = GOODS.ID 
    WHERE GOODS.GOODS_NAME =? AND STOCK.STOCK_TYPE = 0
    '''
    cur.execute(sql_out, (goods_name,))
    total_out = cur.fetchone()[0] or 0

    # 关闭连接
    cur.close()
    con.close()

    return total_in, total_out

def downloadInStockDetail(filename, title, data):

    field_order = ['CATAGORY_NAME', 'GOODS_NAME', 'STOCK_DATE', 'GOODS_NUM',
                   'GOODS_UNIT', 'GOODS_PRICE', 'OP_AREA', 'OP_PERSON', 'TOTAL']

    wb = openpyxl.load_workbook('template/InStockTemplate.xlsx')
    sheet = wb.active
    sheet['A1'].value = title

    row_num = sheet.max_row
    col_end = sheet.max_column

    catagory_start = sheet.max_row+1
    catagory_count = 0
    pre_catagory = None

    goods_start = sheet.max_row+1
    goods_count = 0
    pre_goods = None

    for row in data:
        row_num += 1
        for col in range(col_end):
            field_name = field_order[col]
            cell = sheet.cell(row_num, col+1)
            # 设置单元格边框
            cell.border = border
            if field_name in row.keys():
                # 设置单元格的值
                cell.value = row[field_name]
            if field_name == 'TOTAL':
                cell.value = row['GOODS_NUM']

        catagory_name = row['CATAGORY_NAME']
        goods_name = row['GOODS_NAME']

        # ------------合并物品类别单元格----start------------
        if pre_catagory is None:
            pre_catagory = catagory_name
        if pre_catagory != catagory_name:
            if catagory_count > 1:
                # 当前的类别与上一条记录的类别不一致，且同一类别出现次数大于1时，合并单元格
                c = field_order.index('CATAGORY_NAME')+1
                start_row = catagory_start
                end_row = catagory_start + catagory_count - 1
                # 合并单元格
                sheet.merge_cells(start_row=start_row,
                                  start_column=c, end_row=end_row, end_column=c)
                # 居中对齐
                sheet.cell(start_row, c).alignment = align

            catagory_start += catagory_count
            catagory_count = 0
        catagory_count += 1
        pre_catagory = catagory_name
        # ------------合并物品类别单元格----end------------
        # ------------合并物品单元格 和统计入库笔数----start------------
        if pre_goods is None:
            pre_goods = goods_name
        if pre_goods != goods_name:
            if goods_count > 1:
                # 当前的物品与上一条记录的类别不一致，且同一类别出现次数大于1时，合并单元格
                c = field_order.index('GOODS_NAME')+1
                start_row = goods_start
                end_row = goods_start + goods_count - 1
                # 合并单元格
                sheet.merge_cells(start_row=start_row,
                                  start_column=c, end_row=end_row, end_column=c)
                # 居中对齐
                sheet.cell(start_row, c).alignment = align

                # 统计入库笔数
                # 判断是否是同一规格的物品，若是则计算出总笔数，否则填充黄颜色背景
                c = field_order.index('GOODS_UNIT') + 1
                s = {sheet.cell(r, c).value for r in range(
                    start_row, end_row+1)}
                c_total = field_order.index('TOTAL') + 1
                if len(s) != 1:
                    # 存在不同的规格物品
                    for r in range(start_row, end_row+1):
                        sheet.cell(r, c_total).fill = fill
                else:
                    # 合并计数
                    total = sum(
                        [Fraction(sheet.cell(r, c_total).value) for r in range(start_row, end_row+1)])
                    for r in range(start_row, end_row+1):
                        if total.denominator == 1:
                            sheet.cell(r, c_total).value = str(total)

                        else:
                            total1 = str(int(total)) + ' +' + str(
                                total.numerator - int(total) * total.denominator) + '/' + \
                                     str(total.denominator)

                            sheet.cell(r, c_total).value = total1

                    sheet.merge_cells(start_row=start_row,
                                      start_column=c_total, end_row=end_row, end_column=c_total)

            goods_start += goods_count
            goods_count = 0
        goods_count += 1
        pre_goods = goods_name
        # ------------合并物品单元格----end------------

    if catagory_count > 1:
        c = field_order.index('CATAGORY_NAME')+1
        start_row = catagory_start
        end_row = catagory_start + catagory_count - 1
        sheet.merge_cells(start_row=start_row, start_column=c,
                          end_row=end_row, end_column=c)
        sheet.cell(start_row, c).alignment = align
    if goods_count > 1:
        c = field_order.index('GOODS_NAME')+1
        start_row = goods_start
        end_row = goods_start + goods_count - 1
        sheet.merge_cells(start_row=start_row, start_column=c,
                          end_row=end_row, end_column=c)
        sheet.cell(start_row, c).alignment = align

        # 合并计数
        # 统计入库笔数
        # 判断是否是同一规格的物品，若是则计算出总笔数，否则填充黄颜色背景
        c = field_order.index('GOODS_UNIT') + 1
        s = {sheet.cell(r, c).value for r in range(
            start_row, end_row+1)}
        c_total = field_order.index('TOTAL') + 1
        if len(s) != 1:
            # 存在不同的规格物品
            for r in range(start_row, end_row+1):
                sheet.cell(r, c_total).fill = fill
        else:
            # 合并计数
            total = sum(
                [Fraction(sheet.cell(r, c_total).value) for r in range(start_row, end_row+1)])
            for r in range(start_row, end_row+1):
                if total.denominator == 1:
                    sheet.cell(r, c_total).value = str(total)

                else:
                    total1 = str(int(total)) + ' +' + str(
                        total.numerator - int(total) * total.denominator) + '/' + \
                             str(total.denominator)
                    sheet.cell(r, c_total).value = total1
            sheet.merge_cells(start_row=start_row,
                              start_column=c_total, end_row=end_row, end_column=c_total)

    wb.save(filename)


def downloadOutStockDetail(filename, title, data):

    field_order = ['CATAGORY_NAME', 'GOODS_NAME', 'STOCK_DATE', 'GOODS_NUM',
                   'GOODS_UNIT', 'GOODS_PRICE', 'OP_AREA', 'OP_PERSON', 'TOTAL']

    wb = openpyxl.load_workbook('template/OutStockTemplate.xlsx')
    sheet = wb.active
    sheet['A1'].value = title

    row_num = sheet.max_row
    col_end = sheet.max_column

    catagory_start = sheet.max_row+1
    catagory_count = 0
    pre_catagory = None

    goods_start = sheet.max_row+1
    goods_count = 0
    pre_goods = None

    for row in data:
        row_num += 1
        for col in range(col_end):
            field_name = field_order[col]
            cell = sheet.cell(row_num, col+1)
            # 设置单元格边框
            cell.border = border
            if field_name in row.keys():
                # 设置单元格的值
                cell.value = row[field_name]
            if field_name == 'TOTAL':
                cell.value = row['GOODS_NUM']

        catagory_name = row['CATAGORY_NAME']
        goods_name = row['GOODS_NAME']

        # ------------合并物品类别单元格----start------------
        if pre_catagory is None:
            pre_catagory = catagory_name
        if pre_catagory != catagory_name:
            if catagory_count > 1:
                # 当前的类别与上一条记录的类别不一致，且同一类别出现次数大于1时，合并单元格
                c = field_order.index('CATAGORY_NAME')+1
                start_row = catagory_start
                end_row = catagory_start + catagory_count - 1
                # 合并单元格
                sheet.merge_cells(start_row=start_row,
                                  start_column=c, end_row=end_row, end_column=c)
                # 居中对齐
                sheet.cell(start_row, c).alignment = align

            catagory_start += catagory_count
            catagory_count = 0
        catagory_count += 1
        pre_catagory = catagory_name
        # ------------合并物品类别单元格----end------------
        # ------------合并物品单元格 和统计入库笔数----start------------
        if pre_goods is None:
            pre_goods = goods_name
        if pre_goods != goods_name:
            if goods_count > 1:
                # 当前的物品与上一条记录的类别不一致，且同一类别出现次数大于1时，合并单元格
                c = field_order.index('GOODS_NAME')+1
                start_row = goods_start
                end_row = goods_start + goods_count - 1
                # 合并单元格
                sheet.merge_cells(start_row=start_row,
                                  start_column=c, end_row=end_row, end_column=c)
                # 居中对齐
                sheet.cell(start_row, c).alignment = align

                # 统计入库笔数
                # 判断是否是同一规格的物品，若是则计算出总笔数，否则填充黄颜色背景
                c = field_order.index('GOODS_UNIT') + 1
                s = {sheet.cell(r, c).value for r in range(
                    start_row, end_row+1)}
                c_total = field_order.index('TOTAL') + 1
                if len(s) != 1:
                    # 存在不同的规格物品
                    for r in range(start_row, end_row+1):
                        sheet.cell(r, c_total).fill = fill
                else:
                    # 合并计数
                    total = sum(
                        [Fraction(sheet.cell(r, c_total).value) for r in range(start_row, end_row+1)])
                    for r in range(start_row, end_row+1):
                        if total.denominator == 1:
                            sheet.cell(r, c_total).value = str(total)

                        else:
                            total1 = str(int(total)) + ' +' + str(
                                total.numerator - int(total) * total.denominator) + '/' + \
                                     str(total.denominator)
                            sheet.cell(r, c_total).value = total1
                    sheet.merge_cells(start_row=start_row,
                                      start_column=c_total, end_row=end_row, end_column=c_total)

            goods_start += goods_count
            goods_count = 0
        goods_count += 1
        pre_goods = goods_name
        # ------------合并物品单元格----end------------

    if catagory_count > 1:
        c = field_order.index('CATAGORY_NAME')+1
        start_row = catagory_start
        end_row = catagory_start + catagory_count - 1
        sheet.merge_cells(start_row=start_row, start_column=c,
                          end_row=end_row, end_column=c)
        sheet.cell(start_row, c).alignment = align
    if goods_count > 1:
        c = field_order.index('GOODS_NAME')+1
        start_row = goods_start
        end_row = goods_start + goods_count - 1
        sheet.merge_cells(start_row=start_row, start_column=c,
                          end_row=end_row, end_column=c)
        sheet.cell(start_row, c).alignment = align

        # 合并计数
        # 统计入库笔数
        # 判断是否是同一规格的物品，若是则计算出总笔数，否则填充黄颜色背景
        c = field_order.index('GOODS_UNIT') + 1
        s = {sheet.cell(r, c).value for r in range(
            start_row, end_row+1)}
        c_total = field_order.index('TOTAL') + 1
        if len(s) != 1:
            # 存在不同的规格物品
            for r in range(start_row, end_row+1):
                sheet.cell(r, c_total).fill = fill
        else:
            # 合并计数
            total = sum(
                [Fraction(sheet.cell(r, c_total).value) for r in range(start_row, end_row+1)])
            for r in range(start_row, end_row+1):
                if total.denominator == 1:
                    sheet.cell(r, c_total).value = str(total)

                else:
                    total1 = str(int(total)) + ' +' + str(
                        total.numerator - int(total) * total.denominator) + '/' + \
                             str(total.denominator)
                    sheet.cell(r, c_total).value = total1
            sheet.merge_cells(start_row=start_row,
                              start_column=c_total, end_row=end_row, end_column=c_total)

    wb.save(filename)


def downloadAllStockDetail(filename, title, data):

    # 数据预处理
    catagories = defaultdict(list)
    for row in data:
        catagories[row['CATAGORY_NAME']]

    

    wb = openpyxl.load_workbook('template/StockTemplate.xlsx')
    sheet = wb.active
    sheet['A1'].value = title

    row_count = sheet.max_row

    instock_start= sheet.max_row
    instock_num = 0
    outstock_start= sheet.max_row
    outstock_num = 0
    pre_goodsanme = None

    # 填充数据
    for row in data:
        row_count +=1
        stock_type = row['STOCK_TYPE']
        goodsname = row['GOODS_NAME']
        
        if pre_goodsanme is None:
            pre_goodsanme = goodsname
        
        
        if stock_type == 1:
            if pre_goodsanme != goodsname:
                instock_start += instock_num
                if outstock_num > instock_num:
                    instock_start += outstock_num
                    
                outstock_start = instock_start
                outstock_num = 0    
                instock_num = 0
            instock_num += 1
            for key, val in row.items():
                if key in instock_field_order.keys():
                    col = instock_field_order[key]
                    sheet.cell(instock_start+instock_num, col).value = val
                
                
        elif stock_type == 0:
                
            # 判断是否为同一个物品
            if pre_goodsanme != goodsname:
                outstock_start += outstock_num
                if instock_num > outstock_num:
                    outstock_start += instock_num
                instock_start = outstock_start
                instock_num = 0
                outstock_num = 0
            outstock_num +=1
            for key, val in row.items():
                if key in outstock_field_order.keys():
                    col = outstock_field_order[key]
                    sheet.cell(outstock_start+outstock_num, col).value = val
                    
    # 出入库合计和结余
    c_goodsname = instock_field_order['GOODS_NAME']
    begin = 4;
    end = sheet.max_row
    pre_goods = None
    goods_coordinate = defaultdict(tuple)   
    for r in range(begin, end +1):
        goodsname = sheet.cell(r, c_goodsname).value
        if pre_goods is None:
            pre_goods = goodsname
        if goodsname != pre_goods:
            goods_coordinate[pre_goods] = (begin, r-1)
            begin = r
        pre_goods = goodsname
    goods_coordinate[pre_goods] = (begin, end)            
    #print(goods_coordinate)   
    
    for v in goods_coordinate.values():
        row_start, row_end = v
        caculate_total(sheet,row_start, row_end )
            
    wb.save(filename)

def caculate_total(sheet, row_start, row_end):
    # 入库规格所在列
    unit_col = instock_field_order['GOODS_UNIT']
    # 入库统计所在列
    total_col = instock_field_order['TOTAL']
    # 入库物品数量
    goodsnum_col = instock_field_order['GOODS_NUM']

    in_total = None

    # 判断入库规格是否一致
    unit_num = len({sheet.cell(row, unit_col).value for row in range(row_start, row_end + 1) if sheet.cell(row, unit_col).value is not None})
    if unit_num > 1:
        # 入库规格不一致
        for row in range(row_start, row_end + 1):
            sheet.cell(row, total_col).value = sheet.cell(row, goodsnum_col).value
            sheet.cell(row, total_col).fill = fill
    else:
        total = sum([Fraction(sheet.cell(row, goodsnum_col).value) for row in range(row_start, row_end + 1) if sheet.cell(row, goodsnum_col).value is not None])
        if total.denominator == 1:
            sheet.cell(row_start, total_col).value = str(total)
        else:
            total1 = str(int(total)) + ' +' + str(total.numerator - int(total) * total.denominator) + '/' + str(total.denominator)
            sheet.cell(row_start, total_col).value = total1

        sheet.merge_cells(start_row=row_start, start_column=total_col, end_row=row_end, end_column=total_col)
        in_total = total

    # 出库规格所在列
    unit_col = outstock_field_order['GOODS_UNIT']
    # 出库统计所在列
    total_col = outstock_field_order['TOTAL']
    # 出库物品数量
    goodsnum_col = outstock_field_order['GOODS_NUM']

    out_total = None

    # 判断出库规格是否一致
    unit_num = len({sheet.cell(row, unit_col).value for row in range(row_start, row_end + 1) if sheet.cell(row, unit_col).value is not None})
    if unit_num > 1:
        # 出库规格不一致
        for row in range(row_start, row_end + 1):
            sheet.cell(row, total_col).value = sheet.cell(row, goodsnum_col).value
            sheet.cell(row, total_col).fill = fill
    else:
        total = sum([Fraction(sheet.cell(row, goodsnum_col).value) for row in range(row_start, row_end + 1) if sheet.cell(row, goodsnum_col).value is not None])
        if total.denominator == 1:
            sheet.cell(row_start, total_col).value = str(total)
        else:
            total1 = str(int(total)) + ' +' + str(total.numerator - int(total) * total.denominator) + '/' + str(total.denominator)
            sheet.cell(row_start, total_col).value = total1

        sheet.merge_cells(start_row=row_start, start_column=total_col, end_row=row_end, end_column=total_col)
        out_total = total

    # 获取总库存的出入库数量
    goodsname_col = instock_field_order['GOODS_NAME']
    goods_name = sheet.cell(row_start, goodsname_col).value
    total_in, total_out = get_total_in_out(goods_name)
    total_in = Fraction(total_in)
    total_out = Fraction(total_out)

    # 结余
    c = sheet.max_column
    all_total = total_in - total_out
    if all_total.denominator == 1:
        sheet.cell(row_start, c).value = str(all_total)
    else:
        total1 = str(int(all_total)) + ' +' + str(all_total.numerator - int(all_total) * all_total.denominator) + '/' + str(all_total.denominator)
        sheet.cell(row_start, c).value = total1

    sheet.merge_cells(start_row=row_start, start_column=c, end_row=row_end, end_column=c)